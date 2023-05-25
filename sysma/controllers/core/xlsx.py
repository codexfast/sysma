"""
Read and write xlsx files
"""

import openpyxl
import datetime
import os
import typing

from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import Color
from openpyxl import Workbook

my_red = Color(rgb='00FF0000')
my_green = Color(rgb='0050C878')

def format_with_date(text: str) -> str:
    return datetime.datetime.now().strftime(f"%d-%m-%Y.%H-%M-%S.{text}.xlsx")


class Xlsx:

    """
    xlsx_file is path to Excel file
    """

    def __init__(self, xlsx_file):
        self.xlsx_file = xlsx_file
        self.ss = openpyxl.load_workbook(filename=xlsx_file)
        self.ws = self.ss[self.ss.active.title]


class DataExport:

    def __init__(self, path):
        self.wb = Workbook()
        
        self.sheet = self.wb.active
        self.sheet.title = "SYSPL v1"

        self.path = path

        self.columns = []
        self.row_values = []

    def set_columns(self, columns: list):
        self.columns = columns

    def insert_row(self, values: list):
        self.row_values.append(values)

    def __write(self):
        data = (
            self.columns,
            *self.row_values
        )

        for row in data:
            self.sheet.append(row)


    def __enter__(self):
        return self

    def __exit__(self, type, value, traceback):

        self.__write()

        filename = format_with_date("SYSPL.v2")

        self.wb.save(os.path.join(self.path, filename))

        return False

    